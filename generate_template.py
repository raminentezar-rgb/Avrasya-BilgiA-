# -*- coding: utf-8 -*-
"""
Avrasya Üniversitesi - BilgiAğı (StudyBud)
Toplu Öğrenci Kaydı, Not ve Yoklama Aktarımı için Excel Şablonu Üretici
(Proliz OBS Bağımlılığını Kaldıran Bağımsız Matris Şablonu)
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_student_excel_template(file_path="Avrasya_Ogrenci_ve_Not_Sablonu.xlsx"):
    """
    Öğretim görevlilerinin ve yöneticilerin Proliz OBS'ye ihtiyaç duymadan
    öğrencileri topluca sisteme eklemelerini, odalara kaydetmelerini ve
    başlangıç not/yoklama verilerini yüklemelerini sağlayan Excel şablonunu oluşturur.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Öğrenci & Not Listesi"

    # Avrasya Üniversitesi Renk Paleti ve Stil Tanımları
    navy_fill = PatternFill(start_color="0F3B8C", end_color="0F3B8C", fill_type="solid")
    gold_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    sample_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sample_font = Font(name="Calibri", size=10, bold=False, color="1E293B")
    hint_font = Font(name="Calibri", size=9, italic=True, color="64748B")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Başlık Bilgi Satırı (Ders & Şablon Bilgisi)
    ws.merge_cells("A1:G1")
    ws["A1"] = "AVRASYA ÜNİVERSİTESİ - TOPLU ÖĞRENCİ KAYDI, NOT VE YOKLAMA AKTARIM ŞABLONU"
    ws["A1"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 2. Açıklama Satırı
    ws.merge_cells("A2:G2")
    ws["A2"] = "NOT: İlk 2 satırdaki örnek verileri silip kendi öğrenci listenizi yapıştırınız. Şifre boş bırakılırsa varsayılan olarak 'Avrasya2026!' atanır."
    ws["A2"].font = hint_font
    ws["A2"].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 24

    # 3. Sütun Başlıkları (Row 3)
    headers = [
        "Öğrenci Numarası (Zorunlu)",
        "Adı Soyadı (Zorunlu)",
        "E-Posta Adresi (Zorunlu)",
        "Bölüm / Program",
        "Şifre (Opsiyonel - Boşsa: Avrasya2026!)",
        "Başlangıç Notu / Puan (Opsiyonel)",
        "Katılım / Yoklama Sayısı (Opsiyonel)"
    ]

    ws.row_dimensions[3].height = 28
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        # İlk 3 sütun zorunlu olduğu için Lacivert, diğerleri Opsiyonel olduğu için Altın/Amber
        if col_idx <= 3:
            cell.fill = navy_fill
        else:
            cell.fill = gold_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 4. Örnek Veri Satırları (Row 4 & 5)
    sample_data = [
        ["2024001", "Ahmet Demir", "2024001@ogrenci.avrasya.edu.tr", "Bilgisayar Mühendisliği", "Avrasya2026!", 85, 12],
        ["2024002", "Zeynep Yılmaz", "2024002@ogrenci.avrasya.edu.tr", "Web Tasarım ve Kodlama", "Avrasya2026!", 92, 14],
        ["211101001", "Mehmet Kaya", "211101001@ogrenci.avrasya.edu.tr", "Elektrik-Elektronik Mühendisliği", "", 78, 10]
    ]

    for row_offset, row_data in enumerate(sample_data, start=4):
        ws.row_dimensions[row_offset].height = 22
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.font = sample_font
            cell.fill = sample_fill
            cell.border = thin_border
            if col_idx in [1, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 5. Sütun Genişliklerini Ayarlama
    column_widths = {
        'A': 26,  # Öğrenci No
        'B': 28,  # Adı Soyadı
        'C': 34,  # E-Posta Adresi
        'D': 32,  # Bölüm
        'E': 34,  # Şifre
        'F': 32,  # Başlangıç Notu
        'G': 32   # Yoklama Sayısı
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Dosyayı kaydet
    wb.save(file_path)
    print(f"Başarılı: Excel şablonu '{file_path}' konumuna kaydedildi.")
    return file_path

if __name__ == "__main__":
    generate_student_excel_template()
