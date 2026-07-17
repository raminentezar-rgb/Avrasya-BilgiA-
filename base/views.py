from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib.auth import authenticate, login, logout
from django.utils import timezone
import json
import csv
import socket
import pyotp
from .models import Room, Topic, Message, User, Resource, Quiz, Question, QuizSubmission, QuizSubmissionAnswer, QuestionBankItem, Notification, Assignment, AssignmentSubmission, AttendanceSession, AttendanceRecord, DEPARTMENT_CHOICES
from .forms import RoomForm, UserForm, MyUserCreationForm, ResourceForm, QuizForm, QuestionForm, QuestionBankItemForm
from .proliz_obs import ProlizOBSClient

# Create your views here.

def loginPage(request):
    page = 'login'
    next_url = request.POST.get('next') or request.GET.get('next')
    
    if request.user.is_authenticated:
        if next_url and next_url.startswith('/'):
            return redirect(next_url)
        return redirect('home')

    if request.method == 'POST':
        login_type = request.POST.get('login_type', 'teacher')

        # 1. Proliz OBS Entegrasyonu ile Öğrenci Girişi (Öğrenci Numarası ile)
        if login_type == 'student_obs' or request.POST.get('student_no'):
            student_no = request.POST.get('student_no', '').strip()
            obs_pass = request.POST.get('obs_password', '')
            res = ProlizOBSClient.get_student_info(student_no, obs_pass)
            if res.get('success'):
                email = res['email']
                user = User.objects.filter(Q(student_id=student_no) | Q(email=email)).first()
                if not user:
                    username_base = f"ogr_{student_no}"
                    user = User.objects.create_user(
                        email=email,
                        username=username_base,
                        password=obs_pass or 'AvrasyaOBS123',
                        name=res.get('full_name', f"Öğrenci {student_no}"),
                        student_id=student_no,
                        department=res.get('department', 'Bilgisayar Mühendisliği'),
                        role='student'
                    )
                login(request, user)
                messages.success(request, f"Proliz OBS Entegrasyonu ile hoş geldiniz, {user.name} ({user.department})!")
                if next_url and next_url.startswith('/'):
                    return redirect(next_url)
                return redirect('home')
            else:
                messages.error(request, res.get('error', 'Proliz OBS öğrenci doğrulaması başarısız! Öğrenci Numaranızı kontrol ediniz.'))

        # 2. Akademisyen / Öğretim Üyesi Girişi (Sadece @avrasya.edu.tr e-posta ile)
        else:
            email = request.POST.get('email', '').lower().strip()
            password = request.POST.get('password')

            if not email.endswith('@avrasya.edu.tr'):
                messages.error(request, 'Akademisyen / Öğretim Üyesi girişi için yalnızca resmi Avrasya Üniversitesi e-posta adresi (@avrasya.edu.tr) kullanılmalıdır!')
                return render(request, 'base/login_register.html', {'page': page})

            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Akademisyen olarak giriş yapıldı: {user.name or user.username}")
                if next_url and next_url.startswith('/'):
                    return redirect(next_url)
                return redirect('home')
            else:
                messages.error(request, 'E-posta veya şifre hatalı! Lütfen resmi @avrasya.edu.tr bilgilerinizi kontrol ediniz.')

    context = {'page': page}
    return render(request, 'base/login_register.html', context)


def logoutUser(request):
    logout(request)
    return redirect('home')


def registerPage(request):
    form = MyUserCreationForm()

    if request.method == 'POST':
        form = MyUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            email_clean = user.email.lower().strip() if user.email else ''
            if user.role in ['teacher', 'faculty'] and not email_clean.endswith('@avrasya.edu.tr'):
                messages.error(request, 'Akademisyen / Öğretim Üyesi kaydı için yalnızca resmi Avrasya Üniversitesi e-posta adresi (@avrasya.edu.tr) kullanılmalıdır!')
                return render(request, 'base/login_register.html', {'form': form})
            elif user.role == 'student' and not (email_clean.endswith('@avrasya.edu.tr') or email_clean.endswith('@ogrenci.avrasya.edu.tr')):
                messages.error(request, 'Öğrenci kaydı için yalnızca resmi Avrasya Üniversitesi e-posta adresi (örn: öğrenciadı@avrasya.edu.tr veya @ogrenci.avrasya.edu.tr) kullanılmalıdır. Farklı format kabul edilmez!')
                return render(request, 'base/login_register.html', {'form': form})
            elif not (email_clean.endswith('@avrasya.edu.tr') or email_clean.endswith('@ogrenci.avrasya.edu.tr')):
                messages.error(request, 'Platform kaydı için yalnızca resmi Avrasya Üniversitesi e-posta adresi kullanılmalıdır!')
                return render(request, 'base/login_register.html', {'form': form})
            user.save()
            login(request, user)
            messages.success(request, 'Avrasya BilgiAğı platformuna kaydınız başarıyla tamamlandı!')
            return redirect('home')
        else:
            messages.error(request, 'Kayıt sırasında hata oluştu. Lütfen bilgilerinizi kontrol ediniz.')

    return render(request, 'base/login_register.html', {'form': form})


@login_required(login_url='login')
def home(request):
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    dept = request.GET.get('dept') if request.GET.get('dept') != None else ''

    rooms = Room.objects.filter(
        Q(topic__name__icontains=q) |
        Q(name__icontains=q) |
        Q(course_code__icontains=q) |
        Q(description__icontains=q) |
        Q(host__username__icontains=q) |
        Q(host__department__icontains=q)
    ).distinct()

    if dept:
        rooms = rooms.filter(
            Q(host__department__icontains=dept) |
            Q(topic__name__icontains=dept) |
            Q(name__icontains=dept) |
            Q(description__icontains=dept) |
            Q(course_code__icontains=dept)
        ).distinct()

    # Öğrenciler ve Öğretim Üyeleri (sistem yöneticileri hariç) SADECE kendi bölümlerine ait odaları, konuları ve materyalleri görebilir
    topics = Topic.objects.all()
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        dept = request.user.department
        dept_keywords = [w for w in dept.split() if len(w) > 3 and w not in ['Mühendisliği', 'Fakültesi', 'Bölümü', 'Programı', 'Teknolojisi', 'Hizmetleri']]
        query = Q(host__department=dept) | Q(topic__name__iexact=dept)
        for kw in dept_keywords:
            query |= Q(topic__name__icontains=kw) | Q(name__icontains=kw) | Q(course_code__icontains=kw)
        rooms = rooms.filter(query).distinct()

        # Sol menüdeki konuları da kullanıcının görebildiği odaların konuları veya kendi bölümüyle sınırla
        topics = Topic.objects.filter(
            Q(room__in=rooms) | Q(name__iexact=dept)
        ).distinct()

    room_count = rooms.count()

    room_messages_query = Message.objects.filter(
        Q(room__topic__name__icontains=q))
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        room_messages_query = room_messages_query.filter(room__in=rooms).distinct()
    room_messages = room_messages_query[0:5]

    context = {'rooms': rooms, 'topics': topics,
               'room_count': room_count, 'room_messages': room_messages,
               'departments': DEPARTMENT_CHOICES, 'selected_dept': dept}
    return render(request, 'base/home.html', context)


@login_required(login_url='login')
def dashboard(request):
    user = request.user
    is_faculty = (user.role == 'faculty' or Room.objects.filter(host=user).exists())

    # --- FACULTY DATA ---
    hosted_rooms = Room.objects.filter(host=user).order_by('-updated')
    question_bank_count = QuestionBankItem.objects.filter(creator=user).count()
    recent_quizzes = Quiz.objects.filter(room__host=user).select_related('room').order_by('-created')[:8]
    
    # Pending Quiz essay submissions waiting for grading by this teacher
    pending_quiz_submissions = QuizSubmission.objects.filter(
        Q(quiz__room__host=user) | Q(quiz__creator=user), 
        is_graded=False
    ).select_related('quiz', 'quiz__room', 'student').order_by('-submitted_at')

    # Pending Assignment submissions waiting for grading by this teacher
    pending_ass_submissions = AssignmentSubmission.objects.filter(
        Q(assignment__room__host=user) | Q(assignment__creator=user)
    ).filter(
        Q(grade__isnull=True) | Q(grade__exact='')
    ).select_related('assignment', 'assignment__room', 'student').order_by('-submitted_at')

    # All recent assignment submissions for this teacher (both graded and pending)
    recent_ass_submissions = AssignmentSubmission.objects.filter(
        Q(assignment__room__host=user) | Q(assignment__creator=user)
    ).select_related('assignment', 'assignment__room', 'student').order_by('-submitted_at')[:25]

    # All recent quiz submissions for this teacher (both graded and pending)
    recent_quiz_submissions = QuizSubmission.objects.filter(
        Q(quiz__room__host=user) | Q(quiz__creator=user)
    ).select_related('quiz', 'quiz__room', 'student').order_by('-submitted_at')[:25]

    total_hosted_students = User.objects.filter(participants__host=user).distinct().count()

    # --- STUDENT DATA ---
    enrolled_rooms = Room.objects.filter(participants=user).order_by('-updated')
    my_quiz_submissions = QuizSubmission.objects.filter(student=user).select_related('quiz', 'quiz__room').order_by('-submitted_at')
    my_ass_submissions = AssignmentSubmission.objects.filter(student=user).select_related('assignment', 'assignment__room').order_by('-submitted_at')

    # Active/Upcoming Quizzes across enrolled rooms where student hasn't submitted yet
    submitted_quiz_ids = my_quiz_submissions.values_list('quiz_id', flat=True)
    active_quizzes = Quiz.objects.filter(
        room__participants=user
    ).exclude(id__in=submitted_quiz_ids).select_related('room', 'creator').distinct().order_by('end_time')

    # Active/Upcoming Assignments across enrolled rooms where student hasn't submitted yet
    submitted_ass_ids = my_ass_submissions.values_list('assignment_id', flat=True)
    active_assignments = Assignment.objects.filter(
        room__participants=user
    ).exclude(id__in=submitted_ass_ids).select_related('room', 'creator').distinct().order_by('deadline')

    # Student overall stats
    total_quiz_points_earned = sum(qs.score for qs in my_quiz_submissions)
    total_quiz_points_possible = sum(qs.total_questions for qs in my_quiz_submissions)
    avg_score_pct = int((total_quiz_points_earned / total_quiz_points_possible) * 100) if total_quiz_points_possible > 0 else 0

    # Active attendance sessions across student's enrolled rooms
    active_attendance_sessions = AttendanceSession.objects.filter(
        room__in=enrolled_rooms, is_active=True
    )

    context = {
        'is_faculty': is_faculty,
        'hosted_rooms': hosted_rooms,
        'question_bank_count': question_bank_count,
        'recent_quizzes': recent_quizzes,
        'pending_quiz_submissions': pending_quiz_submissions,
        'pending_ass_submissions': pending_ass_submissions,
        'recent_ass_submissions': recent_ass_submissions,
        'recent_quiz_submissions': recent_quiz_submissions,
        'total_hosted_students': total_hosted_students,
        'enrolled_rooms': enrolled_rooms,
        'my_quiz_submissions': my_quiz_submissions,
        'my_ass_submissions': my_ass_submissions,
        'active_quizzes': active_quizzes,
        'active_assignments': active_assignments,
        'avg_score_pct': avg_score_pct,
        'active_attendance_sessions': active_attendance_sessions,
    }
    return render(request, 'base/dashboard.html', context)


@login_required(login_url='login')
def room(request, pk):
    room_obj = Room.objects.get(id=pk)

    # Öğretim Üyeleri ve Öğrencilerin diğer bölüm odalarına yetkisiz erişimini engelle
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        room_dept = room_obj.topic.name if room_obj.topic else (room_obj.host.department if room_obj.host else '')
        if room_dept and room_dept not in [request.user.department, 'Genel / Diğer'] and (not room_obj.host or room_obj.host.department not in [request.user.department, 'Genel / Diğer']):
            if request.user != room_obj.host and request.user not in room_obj.participants.all():
                messages.error(request, f"Güvenlik Kısıtlaması: Bu oda '{room_dept}' bölümüne aittir. Sadece kendi bölümünüze ({request.user.department}) ait çalışma odalarını görüntüleyebilir veya erişebilirsiniz.")
                return redirect('home')

    room_messages = room_obj.message_set.all()
    participants = room_obj.participants.all()
    resources = room_obj.resources.all()

    resource_form = ResourceForm()

    if request.method == 'POST':
        if 'resource_submit' in request.POST:
            if not request.user.is_authenticated:
                return redirect('login')
            form = ResourceForm(request.POST, request.FILES)
            if form.is_valid():
                res = form.save(commit=False)
                res.user = request.user
                res.room = room_obj
                res.save()
                messages.success(request, 'Ders materyali / sesli anons başarıyla eklendi!')
                return redirect('room', pk=room_obj.id)
            else:
                file_obj = request.FILES.get('file')
                title_str = request.POST.get('title') or 'Ders Materyali / Sesli Anons'
                if file_obj or request.POST.get('link'):
                    Resource.objects.create(
                        user=request.user,
                        room=room_obj,
                        title=title_str,
                        file=file_obj,
                        link=request.POST.get('link', ''),
                        description=request.POST.get('description', '')
                    )
                    messages.success(request, 'Ders materyali / sesli anons başarıyla eklendi!')
                    return redirect('room', pk=room_obj.id)
        elif 'body' in request.POST:
            if not request.user.is_authenticated:
                return redirect('login')
            body_text = request.POST.get('body', '').strip()
            if body_text:
                Message.objects.create(
                    user=request.user,
                    room=room_obj,
                    body=body_text
                )
                room_obj.participants.add(request.user)
                if room_obj.host and request.user != room_obj.host:
                    Notification.objects.create(
                        recipient=room_obj.host,
                        sender=request.user,
                        message=f"@{request.user.username} öğrencisi '{room_obj.name}' odasına yeni bir mesaj yazdı: \"{body_text[:35]}...\"",
                        link=f"/room/{room_obj.id}/"
                    )
            return redirect('room', pk=room_obj.id)

    active_attendance_session = AttendanceSession.objects.filter(
        room=room_obj, is_active=True
    ).first()

    context = {'room': room_obj, 'room_messages': room_messages,
               'participants': participants, 'resources': resources,
               'resource_form': resource_form, 'quizzes': room_obj.quizzes.all(),
               'active_attendance_session': active_attendance_session}
    return render(request, 'base/room.html', context)


@login_required(login_url='login')
def userProfile(request, pk):
    user = User.objects.get(id=pk)
    rooms = user.room_set.all()
    room_messages = user.message_set.all()
    topics = Topic.objects.all()
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        dept = request.user.department
        dept_keywords = [w for w in dept.split() if len(w) > 3 and w not in ['Mühendisliği', 'Fakültesi', 'Bölümü', 'Programı', 'Teknolojisi', 'Hizmetleri']]
        query = Q(host__department=dept) | Q(topic__name__iexact=dept)
        for kw in dept_keywords:
            query |= Q(topic__name__icontains=kw) | Q(name__icontains=kw) | Q(course_code__icontains=kw)
        rooms = rooms.filter(query).distinct()
        room_messages = room_messages.filter(room__in=rooms).distinct()
        topics = topics.filter(Q(room__in=rooms) | Q(name__iexact=dept)).distinct()
    context = {'user': user, 'rooms': rooms,
               'room_messages': room_messages, 'topics': topics}
    return render(request, 'base/profile.html', context)


@login_required(login_url='login')
def createRoom(request):
    if request.user.role not in ['faculty', 'teacher'] and not request.user.is_superuser:
        messages.error(request, 'Ders odası oluşturma yetkisi yalnızca Öğretim Üyeleri / Akademisyenlere aittir.')
        return redirect('home')
    form = RoomForm()
    topics = Topic.objects.all()
    if request.method == 'POST':
        if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
            topic_name = request.user.department
        else:
            topic_name = request.POST.get('topic', 'Genel / Diğer')
        topic, created = Topic.objects.get_or_create(name=topic_name)

        Room.objects.create(
            host=request.user,
            topic=topic,
            name=request.POST.get('name'),
            course_code=request.POST.get('course_code', ''),
            room_type=request.POST.get('room_type', 'general'),
            description=request.POST.get('description'),
        )
        return redirect('home')

    context = {'form': form, 'topics': topics}
    return render(request, 'base/room_form.html', context)


@login_required(login_url='login')
def updateRoom(request, pk):
    room = Room.objects.get(id=pk)
    form = RoomForm(instance=room)
    topics = Topic.objects.all()
    if request.user != room.host and not request.user.is_superuser:
        return HttpResponse('Your are not allowed here!!')

    if request.method == 'POST':
        if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
            topic_name = request.user.department
        else:
            topic_name = request.POST.get('topic', 'Genel / Diğer')
        topic, created = Topic.objects.get_or_create(name=topic_name)
        room.name = request.POST.get('name')
        room.course_code = request.POST.get('course_code', '')
        room.room_type = request.POST.get('room_type', 'general')
        room.topic = topic
        room.description = request.POST.get('description')
        room.save()
        return redirect('home')

    context = {'form': form, 'topics': topics, 'room': room}
    return render(request, 'base/room_form.html', context)


@login_required(login_url='login')
def deleteRoom(request, pk):
    room = Room.objects.get(id=pk)

    if request.user != room.host:
        return HttpResponse('Your are not allowed here!!')

    if request.method == 'POST':
        room.delete()
        return redirect('home')
    return render(request, 'base/delete.html', {'obj': room})


@login_required(login_url='login')
def deleteMessage(request, pk):
    message = Message.objects.get(id=pk)

    if request.user != message.user:
        return HttpResponse('Your are not allowed here!!')

    if request.method == 'POST':
        message.delete()
        return redirect('home')
    return render(request, 'base/delete.html', {'obj': message})


@login_required(login_url='login')
def deleteResource(request, pk):
    resource = Resource.objects.get(id=pk)

    if request.user != resource.user and request.user != resource.room.host:
        return HttpResponse('Bu işlemi yapmaya yetkiniz yok!!')

    if request.method == 'POST':
        room_id = resource.room.id
        resource.delete()
        return redirect('room', pk=room_id)
    return render(request, 'base/delete.html', {'obj': resource})


def roomMessagesAjax(request, pk):
    room_obj = Room.objects.get(id=pk)
    if request.method == 'POST' and request.user.is_authenticated:
        try:
            data = json.loads(request.body)
            body_text = data.get('body', '').strip()
        except:
            body_text = request.POST.get('body', '').strip()

        if body_text:
            msg = Message.objects.create(
                user=request.user,
                room=room_obj,
                body=body_text
            )
            room_obj.participants.add(request.user)
            if room_obj.host and request.user != room_obj.host:
                Notification.objects.create(
                    recipient=room_obj.host,
                    sender=request.user,
                    message=f"@{request.user.username} öğrencisi '{room_obj.name}' odasına yeni bir mesaj yazdı: \"{body_text[:35]}...\"",
                    link=f"/room/{room_obj.id}/"
                )
            return JsonResponse({
                'status': 'ok',
                'message': {
                    'id': msg.id,
                    'body': msg.body,
                    'user': msg.user.username,
                    'name': msg.user.name or msg.user.username,
                    'avatar': msg.user.avatar.url if msg.user.avatar else '/static/images/avatar.svg',
                    'created': msg.created.strftime("%H:%M")
                }
            })
        return JsonResponse({'status': 'error', 'msg': 'Boş mesaj gönderilemez'}, status=400)

    # GET messages
    messages_list = []
    for msg in room_obj.message_set.all():
        messages_list.append({
            'id': msg.id,
            'body': msg.body,
            'user': msg.user.username,
            'name': msg.user.name or msg.user.username,
            'avatar': msg.user.avatar.url if msg.user.avatar else '/static/images/avatar.svg',
            'created': msg.created.strftime("%H:%M"),
            'is_owner': (request.user.is_authenticated and request.user == msg.user)
        })
    return JsonResponse({'status': 'ok', 'messages': messages_list})



@login_required(login_url='login')
def updateUser(request):
    user = request.user
    form = UserForm(instance=user)

    if request.method == 'POST':
        form = UserForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            return redirect('user-profile', pk=user.id)

    return render(request, 'base/update-user.html', {'form': form})


@login_required(login_url='login')
def topicsPage(request):
    q = request.GET.get('q') if request.GET.get('q') != None else ''
    topics = Topic.objects.filter(name__icontains=q)
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        dept = request.user.department
        dept_keywords = [w for w in dept.split() if len(w) > 3 and w not in ['Mühendisliği', 'Fakültesi', 'Bölümü', 'Programı', 'Teknolojisi', 'Hizmetleri']]
        query = Q(name__iexact=dept) | Q(room__host__department=dept)
        for kw in dept_keywords:
            query |= Q(name__icontains=kw)
        topics = topics.filter(query).distinct()
    return render(request, 'base/topics.html', {'topics': topics})


@login_required(login_url='login')
def activityPage(request):
    room_messages = Message.objects.all()
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        dept = request.user.department
        dept_keywords = [w for w in dept.split() if len(w) > 3 and w not in ['Mühendisliği', 'Fakültesi', 'Bölümü', 'Programı', 'Teknolojisi', 'Hizmetleri']]
        query = Q(room__host__department=dept) | Q(room__topic__name__iexact=dept)
        for kw in dept_keywords:
            query |= Q(room__topic__name__icontains=kw) | Q(room__name__icontains=kw) | Q(room__course_code__icontains=kw)
        room_messages = room_messages.filter(query).distinct()
    return render(request, 'base/activity.html', {'room_messages': room_messages})


@login_required(login_url='login')
def create_quiz(request, pk):
    if request.user.role == 'student':
        messages.error(request, "Öğrencilerin sınav oluşturma yetkisi bulunmamaktadır.")
        return redirect('quizzes')

    room = Room.objects.get(id=pk)
    if request.user != room.host and request.user.role != 'faculty':
        return HttpResponse("Bu odaya sınav ekleme yetkiniz bulunmamaktadır.")

    form = QuizForm()
    if request.method == 'POST':
        form = QuizForm(request.POST)
        if form.is_valid():
            quiz = form.save(commit=False)
            quiz.room = room
            quiz.creator = request.user
            quiz.save()
            for p in room.participants.all():
                if p != request.user:
                    Notification.objects.create(
                        recipient=p,
                        sender=request.user,
                        message=f"'{room.name}' odasında yeni çevrimiçi sınav yayınlandı: {quiz.title} 📝",
                        link=f"/quiz/{quiz.id}/"
                    )
            messages.success(request, "Sınav başarıyla oluşturuldu. Şimdi soru ekleyebilirsiniz.")
            return redirect('quiz-detail', pk=quiz.id)

    context = {'form': form, 'room': room}
    return render(request, 'base/quiz_form.html', context)


@login_required(login_url='login')
def create_quiz_general(request):
    if request.user.role == 'student':
        messages.error(request, "Öğrencilerin sınav oluşturma yetkisi bulunmamaktadır.")
        return redirect('quizzes')

    if request.method == 'POST':
        room_id = request.POST.get('room_id')
        room = Room.objects.get(id=room_id)
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        duration_minutes = request.POST.get('duration_minutes', 30)
        
        quiz = Quiz.objects.create(
            room=room,
            creator=request.user,
            title=title,
            description=description,
            duration_minutes=duration_minutes
        )
        for p in room.participants.all():
            if p != request.user:
                Notification.objects.create(
                    recipient=p,
                    sender=request.user,
                    message=f"'{room.name}' odasında yeni çevrimiçi sınav yayınlandı: {quiz.title} 📝",
                    link=f"/quiz/{quiz.id}/"
                )
        messages.success(request, "Sınav başarıyla oluşturuldu! Şimdi aşağıdan soruları ekleyebilir ve sınavı yönetebilirsiniz.")
        return redirect('quiz-detail', pk=quiz.id)

    rooms = Room.objects.all()
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        rooms = rooms.filter(Q(host=request.user) | Q(topic__name__iexact=request.user.department) | Q(host__department=request.user.department)).distinct()
    form = QuizForm()
    return render(request, 'base/quiz_form_general.html', {'rooms': rooms, 'form': form})


@login_required(login_url='login')
def quiz_detail(request, pk):
    quiz = Quiz.objects.get(id=pk)

    # Öğretim Üyeleri ve Öğrencilerin diğer bölüm sınavlarına yetkisiz erişimini engelle
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        room_obj = quiz.room
        room_dept = room_obj.topic.name if room_obj.topic else (room_obj.host.department if room_obj.host else '')
        if room_dept and room_dept not in [request.user.department, 'Genel / Diğer'] and (not room_obj.host or room_obj.host.department not in [request.user.department, 'Genel / Diğer']):
            if request.user != quiz.creator and request.user != room_obj.host and request.user not in room_obj.participants.all():
                messages.error(request, f"Güvenlik Kısıtlaması: Bu sınav '{room_dept}' bölümüne aittir. Sadece kendi bölümünüze ({request.user.department}) ait sınavları görüntüleyebilir veya erişebilirsiniz.")
                return redirect('home')

    questions = quiz.questions.all()
    user_submission = QuizSubmission.objects.filter(quiz=quiz, student=request.user).first()
    question_form = QuestionForm()

    # Öğretim Üyesi / Sınav Sahibi Soru Ekleme
    if request.method == 'POST' and 'add_question' in request.POST:
        if request.user != quiz.creator and request.user != quiz.room.host:
            return HttpResponse("Soru ekleme yetkiniz yok.")
        question_form = QuestionForm(request.POST)
        if question_form.is_valid():
            q = question_form.save(commit=False)
            q.quiz = quiz
            q.save()
            # Eğer 'Soru Bankasına da Kaydet' seçeneği işaretlendiyse
            if request.POST.get('save_to_bank'):
                QuestionBankItem.objects.create(
                    creator=request.user,
                    title=f"{quiz.room.name} - {quiz.title}",
                    question_type=q.question_type,
                    text=q.text,
                    option_a=q.option_a,
                    option_b=q.option_b,
                    option_c=q.option_c,
                    option_d=q.option_d,
                    correct_option=q.correct_option,
                    points=q.points
                )
            messages.success(request, "Soru başarıyla sınav listesine eklendi.")
            return redirect('quiz-detail', pk=quiz.id)

    # Öğrenci Sınav Gönderimi (Cevaplama)
    if request.method == 'POST' and 'submit_quiz' in request.POST:
        if user_submission:
            messages.error(request, "Bu sınavı daha önce cevapladınız.")
            return redirect('quiz-detail', pk=quiz.id)

        total_points = sum(q.points for q in questions) if questions.exists() else 0
        if total_points == 0:
            total_points = questions.count()

        cheat_warnings = 0
        try:
            cheat_warnings = int(request.POST.get('cheat_warnings', 0))
        except ValueError:
            cheat_warnings = 0

        has_essay = any(q.question_type == 'essay' for q in questions)
        
        submission = QuizSubmission.objects.create(
            quiz=quiz,
            student=request.user,
            score=0,
            total_questions=total_points,
            cheat_warnings=cheat_warnings,
            is_graded=not has_essay
        )

        total_awarded = 0
        for q in questions:
            if q.question_type == 'multiple_choice':
                ans = request.POST.get(f'question_{q.id}')
                awarded = q.points if (ans and ans == q.correct_option) else 0
                total_awarded += awarded
                QuizSubmissionAnswer.objects.create(
                    submission=submission,
                    question=q,
                    selected_option=ans,
                    awarded_points=awarded
                )
            else:
                essay_ans = request.POST.get(f'question_{q.id}', '').strip()
                QuizSubmissionAnswer.objects.create(
                    submission=submission,
                    question=q,
                    essay_answer=essay_ans,
                    awarded_points=0
                )

        submission.score = total_awarded
        submission.save()

        if quiz.creator and quiz.creator != request.user:
            Notification.objects.create(
                recipient=quiz.creator,
                sender=request.user,
                message=f"@{request.user.username} öğrencisi '{quiz.title}' sınavını tamamladı! Sonuç/Puan: {total_awarded}/{total_points}",
                link=f"/grade-submission/{submission.id}/" if has_essay else f"/quiz/{quiz.id}/"
            )
        if quiz.room.host and quiz.room.host != request.user and quiz.room.host != quiz.creator:
            Notification.objects.create(
                recipient=quiz.room.host,
                sender=request.user,
                message=f"@{request.user.username} öğrencisi '{quiz.title}' sınavını tamamladı! Sonuç/Puan: {total_awarded}/{total_points}",
                link=f"/grade-submission/{submission.id}/" if has_essay else f"/quiz/{quiz.id}/"
            )

        if has_essay:
            messages.success(request, f"Sınav cevaplarınız başarıyla gönderildi! Yazılı/klasik sorular öğretim üyesi tarafından değerlendirildikten sonra nihai puanınız güncellenecektir. (Test Puanı: {total_awarded} / {total_points})")
        else:
            messages.success(request, f"Sınavı tamamladınız. Sonucunuz: {total_awarded} / {total_points} Puan.")
        return redirect('quiz-detail', pk=quiz.id)

    submissions = quiz.submissions.all()
    question_bank = QuestionBankItem.objects.filter(creator=request.user) if (request.user == quiz.creator or request.user == quiz.room.host) else None

    context = {
        'quiz': quiz,
        'questions': questions,
        'question_form': question_form,
        'user_submission': user_submission,
        'submissions': submissions,
        'question_bank': question_bank,
        'is_creator': (request.user == quiz.creator or request.user == quiz.room.host)
    }
    return render(request, 'base/quiz_detail.html', context)


@login_required(login_url='login')
def delete_quiz(request, pk):
    quiz = Quiz.objects.get(id=pk)
    room_id = quiz.room.id
    if request.user != quiz.creator and request.user != quiz.room.host:
        return HttpResponse("Sınav silme yetkiniz bulunmuyor.")
    if request.method == 'POST':
        quiz.delete()
        messages.success(request, "Sınav silindi.")
        return redirect('room', pk=room_id)
    return render(request, 'base/delete.html', {'obj': quiz.title})


@login_required(login_url='login')
def delete_question(request, pk):
    question = Question.objects.get(id=pk)
    quiz_id = question.quiz.id
    if request.user != question.quiz.creator and request.user != question.quiz.room.host:
        return HttpResponse("Soru silme yetkiniz yok.")
    question.delete()
    messages.success(request, "Soru sınavdan kaldırıldı.")
    return redirect('quiz-detail', pk=quiz_id)


@login_required(login_url='login')
def quizzesPage(request):
    quizzes = Quiz.objects.all()
    if not request.user.is_superuser and request.user.department and request.user.department != 'Genel / Diğer':
        dept = request.user.department
        dept_keywords = [w for w in dept.split() if len(w) > 3 and w not in ['Mühendisliği', 'Fakültesi', 'Bölümü', 'Programı', 'Teknolojisi', 'Hizmetleri']]
        query = Q(room__host__department=dept) | Q(room__topic__name__iexact=dept)
        for kw in dept_keywords:
            query |= Q(room__topic__name__icontains=kw) | Q(room__name__icontains=kw) | Q(room__course_code__icontains=kw)
        quizzes = quizzes.filter(query).distinct()
    return render(request, 'base/quizzes.html', {'quizzes': quizzes})


@login_required(login_url='login')
def export_quiz_results(request, pk):
    quiz = Quiz.objects.get(id=pk)
    if request.user != quiz.creator and request.user != quiz.room.host and request.user.role != 'faculty':
        return HttpResponse("Bu sınavın sonuçlarını indirme yetkiniz yoktur.")

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="sinav_sonuclari_quiz_{quiz.id}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Öğrenci Adı Soyadı', 'Kullanıcı Adı', 'Bölümü', 'Puan', 'Toplam Soru', 'Yüzde (%)', 'Teslim Tarihi'])

    for sub in quiz.submissions.all():
        pct = round((sub.score / sub.total_questions) * 100, 1) if sub.total_questions > 0 else 0
        writer.writerow([
            sub.student.name or sub.student.username,
            sub.student.username,
            sub.student.department or '-',
            sub.score,
            sub.total_questions,
            f"%{pct}",
            sub.submitted_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    return response


@login_required(login_url='login')
def verify_message(request, pk):
    message = Message.objects.get(id=pk)
    room_id = message.room.id
    if request.user == message.room.host or request.user.role == 'faculty':
        message.is_verified_answer = not message.is_verified_answer
        message.save()
        status_txt = "onaylandı" if message.is_verified_answer else "onayı kaldırıldı"
        messages.success(request, f"Cevap öğretim üyesi tarafından {status_txt}.")
        # Create notification for message author if verified
        if message.is_verified_answer and message.user != request.user:
            Notification.objects.create(
                recipient=message.user,
                sender=request.user,
                message=f"'{message.room.name}' odasındaki cevabınız Öğretim Üyesi tarafından Doğru Cevap olarak onaylandı! ✅",
                link=f"/room/{room_id}/"
            )
    else:
        messages.error(request, "Bu işlem için öğretim üyesi yetkiniz olmalıdır.")
    return redirect('room', pk=room_id)


@login_required(login_url='login')
def toggle_save_resource(request, pk):
    resource = Resource.objects.get(id=pk)
    if request.user in resource.saved_by.all():
        resource.saved_by.remove(request.user)
        messages.info(request, "Materyal kütüphanenizden çıkarıldı.")
    else:
        resource.saved_by.add(request.user)
        messages.success(request, "Materyal kişisel kütüphanenize kaydedildi! 🔖")
    return redirect(request.META.get('HTTP_REFERER', 'home'))


@login_required(login_url='login')
def saved_resources(request):
    resources = request.user.saved_resources.all()
    return render(request, 'base/saved_resources.html', {'resources': resources})


@login_required(login_url='login')
def create_assignment(request, pk):
    room = Room.objects.get(id=pk)
    if request.user != room.host and request.user.role != 'faculty':
        return HttpResponse("Ödev oluşturma yetkiniz yok.")

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        deadline = request.POST.get('deadline')
        if title and deadline:
            assignment = Assignment.objects.create(
                room=room,
                creator=request.user,
                title=title,
                description=description or '',
                deadline=deadline
            )
            messages.success(request, "Ödev başarıyla oluşturuldu!")
            # Notify room participants
            for p in room.participants.all():
                if p != request.user:
                    Notification.objects.create(
                        recipient=p,
                        sender=request.user,
                        message=f"'{room.name}' odasında yeni ödev eklendi: {title} (Son Teslim: {deadline})",
                        link=f"/room/{room.id}/"
                    )
    return redirect('room', pk=room.id)


@login_required(login_url='login')
def submit_assignment(request, pk):
    assignment = Assignment.objects.get(id=pk)
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        file = request.FILES.get('file')
        sub, created = AssignmentSubmission.objects.get_or_create(
            assignment=assignment,
            student=request.user,
            defaults={'notes': notes, 'file': file}
        )
        if not created:
            sub.notes = notes
            if file:
                sub.file = file
            sub.grade = ''
            sub.teacher_feedback = ''
            sub.save()
        if assignment.creator and assignment.creator != request.user:
            Notification.objects.create(
                recipient=assignment.creator,
                sender=request.user,
                message=f"@{request.user.username} öğrencisi '{assignment.title}' ödevini teslim etti!",
                link=f"/room/{assignment.room.id}/"
            )
        if assignment.room.host and assignment.room.host != request.user and assignment.room.host != assignment.creator:
            Notification.objects.create(
                recipient=assignment.room.host,
                sender=request.user,
                message=f"@{request.user.username} öğrencisi '{assignment.title}' ödevini teslim etti!",
                link=f"/room/{assignment.room.id}/"
            )
        messages.success(request, "Ödeviniz başarıyla teslim edildi! 🎉")
    return redirect('room', pk=assignment.room.id)


@login_required(login_url='login')
def read_notification(request, pk):
    try:
        notif = Notification.objects.get(id=pk, recipient=request.user)
        link = notif.link or '/'
        notif.is_read = True
        notif.save()
    except Notification.DoesNotExist:
        link = '/'
    return redirect(link)


@login_required(login_url='login')
def clear_all_notifications(request):
    request.user.notifications.filter(is_read=False).update(is_read=True)
    return redirect(request.META.get('HTTP_REFERER', 'home'))


@login_required(login_url='login')
def add_from_bank_to_quiz(request, quiz_id, bank_id):
    quiz = Quiz.objects.get(id=quiz_id)
    if request.user != quiz.creator and request.user != quiz.room.host:
        return HttpResponse("Bu işlem için yetkiniz yok.")
    
    try:
        item = QuestionBankItem.objects.get(id=bank_id, creator=request.user)
        Question.objects.create(
            quiz=quiz,
            question_type=item.question_type,
            text=item.text,
            option_a=item.option_a or '',
            option_b=item.option_b or '',
            option_c=item.option_c or '',
            option_d=item.option_d or '',
            correct_option=item.correct_option or 'A',
            points=item.points
        )
        messages.success(request, f"'{item.title}' sorusu bankadan sınavınıza eklendi!")
    except QuestionBankItem.DoesNotExist:
        messages.error(request, "Soru bankasında bu öğe bulunamadı.")
    return redirect('quiz-detail', pk=quiz.id)


@login_required(login_url='login')
def question_bank_view(request):
    if request.user.role not in ['faculty', 'teacher'] and not request.user.is_superuser:
        messages.error(request, "Soru bankasına sadece öğretim üyeleri erişebilir.")
        return redirect('home')

    items = QuestionBankItem.objects.filter(creator=request.user)
    form = QuestionBankItemForm()

    if request.method == 'POST':
        if 'delete_bank_item' in request.POST:
            item_id = request.POST.get('delete_bank_item')
            QuestionBankItem.objects.filter(id=item_id, creator=request.user).delete()
            messages.success(request, "Soru bankasından silindi.")
            return redirect('question-bank')
        
        form = QuestionBankItemForm(request.POST)
        if form.is_valid():
            bank_item = form.save(commit=False)
            bank_item.creator = request.user
            bank_item.save()
            messages.success(request, "Soru başarıyla kişisel bankanıza eklendi.")
            return redirect('question-bank')

    context = {'items': items, 'form': form}
    return render(request, 'base/question_bank.html', context)


@login_required(login_url='login')
def grade_quiz_submission(request, submission_id):
    submission = QuizSubmission.objects.get(id=submission_id)
    quiz = submission.quiz
    if request.user != quiz.creator and request.user != quiz.room.host:
        return HttpResponse("Değerlendirme yetkiniz bulunmuyor.")

    if request.method == 'POST':
        for ans in submission.answers.all():
            if ans.question.question_type == 'essay':
                pts = request.POST.get(f'points_{ans.id}')
                comment = request.POST.get(f'comment_{ans.id}', '')
                if pts is not None:
                    try:
                        ans.awarded_points = int(pts)
                    except ValueError:
                        pass
                ans.teacher_comment = comment
                ans.save()

        total = sum(a.awarded_points for a in submission.answers.all())
        submission.score = total
        submission.teacher_feedback = request.POST.get('teacher_feedback', '')
        submission.is_graded = True
        submission.save()

        Notification.objects.create(
            recipient=submission.student,
            sender=request.user,
            message=f"'{quiz.title}' sınavınızdaki yazılı sorular puanlandı! Notunuz: {total}/{submission.total_questions}",
            link=f"/quiz/{quiz.id}/"
        )
        messages.success(request, f"Öğrencinin sınav puanı ve değerlendirmesi güncellendi! Toplam: {total}/{submission.total_questions}")
        return redirect('grade-quiz-submission', submission_id=submission.id)

    context = {
        'submission': submission,
        'quiz': quiz,
        'answers': submission.answers.all()
    }
    return render(request, 'base/grade_quiz_submission.html', context)


@login_required(login_url='login')
def grade_assignment_submission(request, submission_id):
    sub = AssignmentSubmission.objects.get(id=submission_id)
    room = sub.assignment.room
    if request.user != sub.assignment.creator and request.user != room.host:
        return HttpResponse("Ödev notlandırma yetkiniz yok.")

    if request.method == 'POST':
        grade = request.POST.get('grade', '').strip()
        feedback = request.POST.get('teacher_feedback', '').strip()
        sub.grade = grade
        sub.teacher_feedback = feedback
        sub.save()

        Notification.objects.create(
            recipient=sub.student,
            sender=request.user,
            message=f"'{sub.assignment.title}' ödevinize öğretim üyesi not ve değerlendirme girdi: {grade}",
            link=f"/room/{room.id}/"
        )
        messages.success(request, f"@{sub.student.username} öğrencisinin ödev notu ('{grade}') ve geri bildirimi kaydedildi.")
        return redirect('room', pk=room.id)

    return redirect('room', pk=room.id)


@login_required(login_url='login')
def room_gradebook(request, pk):
    room = Room.objects.get(id=pk)
    if request.user != room.host and request.user.role != 'faculty' and not request.user.is_superuser:
        messages.error(request, "Ders not defterine sadece odanın öğretim üyesi erişebilir.")
        return redirect('room', pk=room.id)

    students = room.participants.all()
    quizzes = room.quizzes.all()
    assignments = room.assignments.all()

    gradebook_rows = []
    for student in students:
        quiz_scores = []
        total_quiz_pts = 0
        max_quiz_pts = 0
        for q in quizzes:
            sub = QuizSubmission.objects.filter(quiz=q, student=student).first()
            if sub:
                quiz_scores.append({'score': sub.score, 'total': sub.total_questions, 'submitted': True, 'is_graded': sub.is_graded, 'sub_id': sub.id})
                total_quiz_pts += sub.score
                max_quiz_pts += sub.total_questions
            else:
                quiz_scores.append({'score': '-', 'total': '-', 'submitted': False})

        ass_scores = []
        for a in assignments:
            sub = AssignmentSubmission.objects.filter(assignment=a, student=student).first()
            if sub:
                ass_scores.append({'grade': sub.grade or 'Not Girilmedi', 'submitted': True, 'sub_id': sub.id, 'file_url': sub.file.url if sub.file else None})
            else:
                ass_scores.append({'grade': 'Teslim Edilmedi', 'submitted': False})

        quiz_percent = round((total_quiz_pts / max_quiz_pts) * 100) if max_quiz_pts > 0 else 0
        gradebook_rows.append({
            'student': student,
            'quiz_scores': quiz_scores,
            'ass_scores': ass_scores,
            'quiz_percent': quiz_percent
        })

    context = {
        'room': room,
        'quizzes': quizzes,
        'assignments': assignments,
        'gradebook_rows': gradebook_rows
    }
    return render(request, 'base/room_gradebook.html', context)


@login_required(login_url='login')
def export_room_gradebook(request, pk):
    room = Room.objects.get(id=pk)
    if request.user != room.host and request.user.role != 'faculty' and not request.user.is_superuser:
        return HttpResponse("Yetkisiz erişim.")

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="room_{room.id}_gradebook.csv"'
    response.write('\ufeff'.encode('utf8')) # BOM for Excel UTF-8 support
    writer = csv.writer(response, delimiter=';')

    students = room.participants.all()
    quizzes = room.quizzes.all()
    assignments = room.assignments.all()

    # Header row
    headers = ['Öğrenci Adı Soyadı', 'Kullanıcı Adı', 'Bölüm / Fakülte']
    for q in quizzes:
        headers.append(f"Sınav: {q.title} (Puan)")
    for a in assignments:
        headers.append(f"Ödev: {a.title} (Not)")
    headers.append("Ortalama Sınav Başarı (%)")
    writer.writerow(headers)

    for student in students:
        row = [student.name or student.username, f"@{student.username}", student.department or "-"]
        total_quiz_pts = 0
        max_quiz_pts = 0
        for q in quizzes:
            sub = QuizSubmission.objects.filter(quiz=q, student=student).first()
            if sub:
                row.append(f"{sub.score} / {sub.total_questions}")
                total_quiz_pts += sub.score
                max_quiz_pts += sub.total_questions
            else:
                row.append("Katılmadı")

        for a in assignments:
            sub = AssignmentSubmission.objects.filter(assignment=a, student=student).first()
            if sub:
                row.append(sub.grade or "Not Girilmedi")
            else:
                row.append("Teslim Edilmedi")

        quiz_percent = round((total_quiz_pts / max_quiz_pts) * 100) if max_quiz_pts > 0 else 0
        row.append(f"%{quiz_percent}")
        writer.writerow(row)

    return response


# ==============================================================================
# CLASSROOM ATTENDANCE SYSTEM WITH DYNAMIC QR CODE & WI-FI IP CHECK
# ==============================================================================

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR', '')
    return ip

def get_local_lan_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

def is_ip_allowed(session_ip, student_ip):
    if not session_ip or not student_ip:
        return True
    if session_ip == student_ip:
        return True
    if student_ip in ('127.0.0.1', '::1', 'localhost') or session_ip in ('127.0.0.1', '::1', 'localhost'):
        return True
    t_parts = session_ip.split('.')
    s_parts = student_ip.split('.')
    if len(t_parts) == 4 and len(s_parts) == 4:
        # Allow /24 exact match or /16 exact match for university classroom Wi-Fi NAT pools and load balancers
        if t_parts[:3] == s_parts[:3] or t_parts[:2] == s_parts[:2]:
            return True
    return False

@login_required
def start_attendance_session(request, room_id):
    room = get_object_or_404(Room, id=room_id)
    # Only room host or superuser can start attendance
    if room.host != request.user and not request.user.is_superuser:
        messages.error(request, "Bu ders odasında yoklama başlatma yetkiniz bulunmuyor.")
        return redirect('room', pk=room.id)
    
    # Prevent opening a new attendance session until Yoklamayi Bitir is clicked
    if not request.user.is_superuser:
        active_session = AttendanceSession.objects.filter(room__host=request.user, is_active=True).first()
    else:
        active_session = AttendanceSession.objects.filter(room=room, is_active=True).first()
        
    if active_session:
        messages.warning(
            request,
            f"'{active_session.room.name}' odasında halen açık/aktif bir yoklama oturumunuz bulunmaktadır! "
            "Yeni bir yoklama listesi açabilmek için öncelikle mevcut ekrandan 'Yoklamayı Bitir (Odaya Dön)' butonuna basmalısınız."
        )
        return redirect('projector_view', session_id=active_session.id)
    
    teacher_ip = get_client_ip(request)
    if teacher_ip in ('127.0.0.1', '::1', 'localhost'):
        teacher_ip = get_local_lan_ip()
        
    session = AttendanceSession.objects.create(
        room=room,
        teacher_ip=teacher_ip,
        require_ip_check=True
    )
    return redirect('projector_view', session_id=session.id)

@login_required
def projector_view(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        messages.info(request, "Bu yoklama oturumu artık mevcut değil (Sunucu güncellenmiş veya oturum kapanmış olabilir).")
        return redirect('home')
    if session.room.host != request.user and not request.user.is_superuser:
        messages.error(request, "Yetkisiz erişim.")
        return redirect('home')
    
    server_lan_ip = get_local_lan_ip()
    context = {'session': session, 'server_lan_ip': server_lan_ip}
    return render(request, 'base/projector_view.html', context)

@login_required
def api_projector_token(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        return JsonResponse({'error': 'Session not found'}, status=404)
    if not session.is_active:
        return JsonResponse({'error': 'Session inactive'}, status=400)
    return JsonResponse({'token': session.get_totp_token()})

@login_required
def api_projector_live(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        return JsonResponse({'error': 'Session not found'}, status=404)
    records = session.records.all().order_by('-timestamp')
    data = []
    for r in records:
        data.append({
            'student_name': f"{r.student.name}" if r.student.name else r.student.username,
            'student_no': r.student.student_id or r.student.username,
            'time': timezone.localtime(r.timestamp).strftime('%H:%M:%S'),
            'ip': r.client_ip or '-'
        })
    return JsonResponse({'records': data, 'count': records.count()})

@login_required
def toggle_ip_check(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        return JsonResponse({'error': 'Session not found'}, status=404)
    if session.room.host == request.user or request.user.is_superuser:
        session.require_ip_check = not session.require_ip_check
        session.save()
        return JsonResponse({'status': True, 'require_ip_check': session.require_ip_check})
    return JsonResponse({'status': False}, status=403)

@login_required
def toggle_qr_check(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        return JsonResponse({'error': 'Session not found'}, status=404)
    if session.room.host == request.user or request.user.is_superuser:
        session.allow_qr_check = not session.allow_qr_check
        session.save()
        return JsonResponse({'status': True, 'allow_qr_check': session.allow_qr_check})
    return JsonResponse({'status': False}, status=403)

@login_required
def close_attendance_session(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        messages.info(request, "Yoklama oturumu artık mevcut değil veya kapatılmış.")
        return redirect('home')
    if session.room.host == request.user or request.user.is_superuser:
        session.is_active = False
        session.save()
        messages.success(request, "Yoklama oturumu başarıyla sonlandırıldı.")
    return redirect('room', pk=session.room.id)

@login_required
def student_scan(request):
    session_id = request.GET.get('session')
    token = request.GET.get('token')
    
    if not session_id or not token:
        messages.error(request, 'Geçersiz yoklama bağlantısı veya QR kod bilgisi eksik.')
        return render(request, 'base/scan_result.html', {'success': False})
        
    student = request.user
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        messages.error(request, 'Yoklama oturumu bulunamadı veya sunucu yenilenmesi nedeniyle sona ermiş. Lütfen projektördeki yeni kodu okutun.')
        return render(request, 'base/scan_result.html', {'success': False})
        
    # Checks
    if not session.is_active:
        messages.warning(request, 'Bu yoklama oturumu sona ermiş.')
        return render(request, 'base/scan_result.html', {'success': False})
        
    if student not in session.room.participants.all() and student != session.room.host:
        # Automatically enroll student when scanning class QR code
        session.room.participants.add(student)
        
    if not session.allow_qr_check:
        messages.error(request, 'Bu oturumda QR Kod ile yoklama kapalıdır. Lütfen Sınıf Wi-Fi Ağına bağlanarak otomatik yoklama verin.')
        return render(request, 'base/scan_result.html', {'success': False})

    if not session.verify_totp(token):
        messages.error(request, 'QR kodun süresi dolmuş. Lütfen tahtadaki/projektördeki yeni kodu okutun.')
        return render(request, 'base/scan_result.html', {'success': False})

    # IP Based Verification
    student_ip = get_client_ip(request)
    if session.require_ip_check and session.teacher_ip:
        if not is_ip_allowed(session.teacher_ip, student_ip):
            messages.error(request, f'IP Koruması Hatası: Sınıf Wi-Fi ağından bağlanmamış görünüyorsunuz! (Sizin IP: {student_ip} | Sınıf IP: {session.teacher_ip}). Lütfen VPN / Hücresel veriyi kapatıp sınıf Wi-Fi ağına bağlanın.')
            return render(request, 'base/scan_result.html', {'success': False, 'error_msg': f'Sizin IP Adresiniz ({student_ip}), Sınıf Wi-Fi Ağı IP Adresi ({session.teacher_ip}) ile eşleşmedi.'})
        
    # Record Checkin
    record, created = AttendanceRecord.objects.get_or_create(
        session=session,
        student=student,
        defaults={'client_ip': student_ip, 'status': 'present'}
    )
    if not created:
        messages.info(request, 'Bu ders için yoklamanız zaten alınmıştı.')
    else:
        messages.success(request, f'Başarıyla yoklamanız alındı! ({session.room.name})')
        
        # Send notification to professor
        if session.room.host and session.room.host != student:
            Notification.objects.create(
                recipient=session.room.host,
                sender=student,
                message=f"📋 {student.name or student.username} yoklamaya katıldı ({session.room.name}).",
                link=f"/room/{session.room.id}/"
            )
        
    return render(request, 'base/scan_result.html', {'success': True, 'room_name': session.room.name})

@login_required
def student_ip_checkin(request):
    session_id = request.GET.get('session')
    if not session_id:
        messages.error(request, 'Geçersiz yoklama oturumu bilgisi.')
        return redirect('dashboard')
        
    student = request.user
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session or not session.is_active:
        messages.warning(request, 'Bu yoklama oturumu artık aktif değil veya sona ermiş.')
        return redirect('dashboard')
        
    # Automatically enroll student inside class room
    if student not in session.room.participants.all() and student != session.room.host:
        session.room.participants.add(student)
        
    # Verify Wi-Fi IP
    student_ip = get_client_ip(request)
    if session.require_ip_check and session.teacher_ip:
        if not is_ip_allowed(session.teacher_ip, student_ip):
            messages.error(request, f'IP Koruması Hatası: Sınıf Wi-Fi ağından bağlanmamış görünüyorsunuz! (Sizin IP: {student_ip} | Sınıf IP: {session.teacher_ip}). Lütfen VPN / Hücresel veriyi kapatıp sınıf Wi-Fi ağına bağlanın.')
            return render(request, 'base/scan_result.html', {'success': False, 'error_msg': f'Sizin IP Adresiniz ({student_ip}), Sınıf Wi-Fi Ağı IP Adresi ({session.teacher_ip}) ile eşleşmedi.'})
            
    # Record Checkin
    record, created = AttendanceRecord.objects.get_or_create(
        session=session,
        student=student,
        defaults={'client_ip': student_ip, 'status': 'present'}
    )
    if not created:
        messages.info(request, 'Bu ders için yoklamanız zaten alınmıştı.')
    else:
        messages.success(request, f'📡 Wi-Fi IP ile yoklamanız başarıyla alındı! ({session.room.name})')
        
        # Send notification to professor
        if session.room.host and session.room.host != student:
            Notification.objects.create(
                recipient=session.room.host,
                sender=student,
                message=f"📡 {student.name or student.username} Wi-Fi IP ile yoklamaya katıldı ({session.room.name}).",
                link=f"/room/{session.room.id}/"
            )
            
    return render(request, 'base/scan_result.html', {'success': True, 'room_name': session.room.name})

@login_required
def export_attendance_report(request, session_id):
    session = AttendanceSession.objects.filter(id=session_id).first()
    if not session:
        messages.info(request, "Rapor oluşturulacak yoklama oturumu bulunamadı.")
        return redirect('home')
    if session.room.host != request.user and not request.user.is_superuser:
        messages.error(request, "Yetkisiz erişim.")
        return redirect('home')
        
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Yoklama Raporu"
    
    # Styling
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    present_font = Font(name="Arial", size=10, bold=True, color="166534")
    absent_font = Font(name="Arial", size=10, bold=True, color="991B1B")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title rows
    ws.merge_cells("A1:F1")
    ws["A1"] = f"AVRASYA BİLGİAĞI - YOKLAMA RAPORU ({session.room.name.upper()})"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    ws["A2"] = f"Ders / Oda: {session.room.name}"
    ws["D2"] = f"Tarih: {session.created_at.strftime('%d.%m.%Y %H:%M')}"
    ws["A3"] = f"Öğretim Üyesi: {session.room.host.name if session.room.host and session.room.host.name else (session.room.host.username if session.room.host else '-')}"
    ws["D3"] = f"Sınıf IP Adresi: {session.teacher_ip or 'Tanımsız'}"
    
    # Table headers
    headers = ["Sıra", "Öğrenci No / Sicil", "Ad Soyad", "Bölüm / Fakülte", "Yoklama Durumu", "Katılım Saati & IP"]
    ws.append([]) # Row 4 blank
    ws.append(headers) # Row 5 headers
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 24
    
    # Get all participants of the room
    participants = session.room.participants.all().exclude(id=session.room.host.id if session.room.host else -1).order_by('username')
    records_map = {r.student.id: r for r in session.records.all()}
    
    row_num = 6
    present_count = 0
    absent_count = 0
    
    for idx, student in enumerate(participants, 1):
        record = records_map.get(student.id)
        is_present = record is not None
        if is_present:
            present_count += 1
            status_str = "VAR (Katıldı)"
            time_ip_str = f"{timezone.localtime(record.timestamp).strftime('%H:%M:%S')} (IP: {record.client_ip or '-'})"
        else:
            absent_count += 1
            status_str = "YOK (Katılmadı)"
            time_ip_str = "-"
            
        row_data = [
            idx,
            student.student_id or student.username,
            student.name or student.username,
            student.department or "-",
            status_str,
            time_ip_str
        ]
        ws.append(row_data)
        
        # Formatting row cells
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if col_idx == 5:
                cell.fill = present_fill if is_present else absent_fill
                cell.font = present_font if is_present else absent_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (1, 2, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 20
        row_num += 1
        
    # Summary row at bottom
    ws.append([])
    ws.append([f"TOPLAM: {len(participants)} Öğrenci | KATILAN (VAR): {present_count} | KATILMAYAN (YOK): {absent_count}"])
    ws.merge_cells(f"A{row_num+1}:F{row_num+1}")
    ws.cell(row=row_num+1, column=1).font = Font(name="Arial", size=11, bold=True, color="1E3A8A")
    
    # Auto-fit columns safely without MergedCell conflicts
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 7):
        col_letter = get_column_letter(col_idx)
        max_len = 15
        for row_idx in range(5, row_num):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
            if len(cell_val) > max_len:
                max_len = len(cell_val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Yoklama_Raporu_{session.room.name}_{session.created_at.strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def download_student_template(request):
    """
    Excel ile Toplu Öğrenci Kaydı, Not ve Yoklama Şablonunu dinamik olarak veya dosyadan indirir.
    """
    import io
    from generate_template import generate_student_excel_template
    
    output = io.BytesIO()
    wb_path = generate_student_excel_template()
    with open(wb_path, 'rb') as f:
        output.write(f.read())
    output.seek(0)
    
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Avrasya_Ogrenci_ve_Not_Sablonu.xlsx"'
    return response


@login_required(login_url='login')
def bulk_import_students(request, room_id=None):
    """
    Öğretim üyeleri tarafından yüklenen Excel veya CSV dosyasını işleyerek:
    1. Öğrenci hesaplarını (User - student) veritabanında oluşturur veya günceller (Proliz API'ye gerek kalmaz)
    2. Seçili odaya (Room) öğrencileri topluca kaydeder
    3. Opsiyonel olarak Başlangıç Notu ve Yoklama sayısını işler.
    """
    if request.method != 'POST':
        return redirect('home')

    room = None
    if room_id:
        room = get_object_or_404(Room, id=room_id)
        if request.user != room.host and request.user.role != 'faculty' and not request.user.is_superuser:
            messages.error(request, "Bu odaya toplu öğrenci ve not ekleme yetkiniz bulunmamaktadır.")
            return redirect('room', pk=room.id)
    elif request.user.role != 'faculty' and not request.user.is_superuser:
        messages.error(request, "Toplu öğrenci ve not ekleme işlemi için öğretim üyesi yetkisi gerekir.")
        return redirect('dashboard')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, "Lütfen yüklemek için bir Excel (.xlsx/.xls) veya CSV dosyası seçiniz!")
        return redirect('room', pk=room.id) if room else redirect('dashboard')

    filename = excel_file.name.lower()
    created_count = 0
    updated_count = 0
    enrolled_count = 0
    graded_count = 0
    attendance_recorded_count = 0

    try:
        rows = []
        if filename.endswith('.csv'):
            import csv
            decoded_file = excel_file.read().decode('utf-8-sig').splitlines()
            reader = csv.reader(decoded_file, delimiter=';')
            rows = list(reader)
            if rows and len(rows[0]) <= 1:
                reader = csv.reader(decoded_file, delimiter=',')
                rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if any(r):
                    rows.append([str(c).strip() if c is not None else '' for c in r])

        # Filter out header/info rows
        data_rows = []
        for r in rows:
            if not r or not r[0]:
                continue
            first_val = str(r[0]).strip().lower()
            if any(kw in first_val for kw in ['avrasya', 'öğrenci no', 'ogrenci no', 'öğrenci numarası', 'not:', 'toplam', 'zorunlu']):
                continue
            data_rows.append(r)

        if not data_rows:
            messages.warning(request, "Yüklenen dosyada geçerli öğrenci veri satırı bulunamadı! Lütfen şablonu kontrol ediniz.")
            return redirect('room', pk=room.id) if room else redirect('dashboard')

        from django.db import transaction
        with transaction.atomic():
            for idx, row in enumerate(data_rows, start=1):
                if len(row) < 2 or not row[0].strip() or not row[1].strip():
                    continue
                student_id = str(row[0]).strip()
                full_name = str(row[1]).strip()
                email = str(row[2]).strip() if len(row) > 2 and row[2].strip() else f"{student_id}@ogrenci.avrasya.edu.tr"
                department = str(row[3]).strip() if len(row) > 3 and row[3].strip() else 'Bilgisayar Mühendisliği'
                password = str(row[4]).strip() if len(row) > 4 and row[4].strip() else 'Avrasya2026!'

                initial_grade = None
                if len(row) > 5 and row[5].strip():
                    try:
                        initial_grade = float(row[5].strip().replace(',', '.'))
                    except ValueError:
                        pass

                attendance_count = None
                if len(row) > 6 and row[6].strip():
                    try:
                        attendance_count = int(float(row[6].strip()))
                    except ValueError:
                        pass

                user = User.objects.filter(Q(student_id=student_id) | Q(email=email)).first()
                if not user:
                    username_base = f"ogr_{student_id}"
                    if User.objects.filter(username=username_base).exists():
                        username_base = f"ogr_{student_id}_{idx}"
                    user = User.objects.create_user(
                        username=username_base,
                        email=email,
                        password=password,
                        name=full_name,
                        student_id=student_id,
                        department=department,
                        role='student'
                    )
                    created_count += 1
                else:
                    changed = False
                    if not user.name or user.name == user.username:
                        user.name = full_name
                        changed = True
                    if not user.student_id:
                        user.student_id = student_id
                        changed = True
                    if department and department != 'Bilgisayar Mühendisliği':
                        user.department = department
                        changed = True
                    if changed:
                        user.save()
                        updated_count += 1

                # Enroll in Room if room provided
                if room:
                    if user not in room.participants.all():
                        room.participants.add(user)
                        enrolled_count += 1

                    # Process grade if provided
                    if initial_grade is not None:
                        quiz, _ = Quiz.objects.get_or_create(
                            room=room,
                            title=f"Excel Aktarılan Notlar ({room.name})",
                            defaults={
                                'description': 'Excel şablonundan toplu yüklenen başlangıç/ara sınav notları',
                                'creator': request.user,
                                'is_active': False
                            }
                        )
                        sub, sub_created = QuizSubmission.objects.get_or_create(
                            quiz=quiz,
                            student=user,
                            defaults={
                                'score': initial_grade,
                                'total_questions': 100,
                                'is_graded': True,
                                'teacher_feedback': f"Excel ile toplu yüklendi (Başlangıç Puanı: {initial_grade})"
                            }
                        )
                        if not sub_created:
                            sub.score = initial_grade
                            sub.is_graded = True
                            sub.teacher_feedback = f"Excel ile güncellendi (Başlangıç Puanı: {initial_grade})"
                            sub.save()
                        graded_count += 1

                    # Process historical attendance count if provided
                    if attendance_count is not None and attendance_count > 0:
                        session, _ = AttendanceSession.objects.get_or_create(
                            room=room,
                            is_active=False,
                            defaults={'teacher_ip': '127.0.0.1'}
                        )
                        rec, rec_created = AttendanceRecord.objects.get_or_create(
                            session=session,
                            student=user,
                            defaults={
                                'status': 'present',
                                'client_ip': f"Excel_Aktarimi ({attendance_count} Katilim)"
                            }
                        )
                        if not rec_created and not rec.client_ip.startswith('Excel_Aktarimi'):
                            rec.client_ip = f"Excel_Aktarimi ({attendance_count} Katilim)"
                            rec.save()
                        attendance_recorded_count += 1

        msg = f"Excel İşlemi Tamamlandı: {created_count} yeni öğrenci hesabı oluşturuldu, {updated_count} hesap güncellendi."
        if room:
            msg += f" {enrolled_count} öğrenci '{room.name}' odasına eklendi."
            if graded_count > 0:
                msg += f" {graded_count} öğrencinin notu Not Defterine ('Excel Aktarılan Notlar') işlendi."
            if attendance_recorded_count > 0:
                msg += f" {attendance_recorded_count} öğrencinin yoklama kaydı aktarıldı."
        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f"Excel dosyası işlenirken bir hata oluştu: {str(e)}")

    return redirect('room', pk=room.id) if room else redirect('dashboard')

